"""
Simple script to programmatically fill fields in RGA Tool Excel file
Focuses only on 'General Information' and 'Benefit Selection - Group' sheets
Works on any operating system with Python and openpyxl
"""

import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def fill_rga_tool(file_path, output_path=None):
    """
    Fill fields in the RGA Tool Excel file with predefined values
    
    Args:
        file_path (str): Path to the Excel file
        output_path (str, optional): Path to save the filled file. If None, creates a new filename.
    
    Returns:
        str: Path to saved file
    """
    if output_path is None:
        # Create output filename with timestamp
        dirname, filename = os.path.split(file_path)
        base, ext = os.path.splitext(filename)
        output_path = os.path.join(dirname, f"{base}_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}")
    
    print(f"Loading workbook: {file_path}")
    
    # Load the workbook with keep_vba=True to preserve macros
    wb = load_workbook(file_path, keep_vba=True)
    
    # Fill General Information sheet
    print("Filling General Information sheet...")
    if "General Information" in wb.sheetnames:
        sheet = wb["General Information"]
        
        # Define the values to fill (cell_address: value)
        general_info_data = {
            'C5': 'HUHU',
            'D7': 'FALAK TAYYEB PLATINUM GOVERNMENT SERVICES LLC',
            'D9': 'New Business',
            'D11': datetime(2024, 4, 14),
            'D13': datetime(2025, 4, 14),
            'D15': datetime(2026, 4, 13),
            'D19': 'No',
            'D21': 'UAE',
            'D23': 100,     # Number of employees
            'D25': 150,     # Number of members
            'D27': 'Insurance Broker XYZ',  # Intermediary name
            'D29': 10,      # Commission %
            'D31': 'John Smith',  # Contact Person
            'D33': 'john.smith@example.com',  # Email
            'D35': '+971 50 123 4567',  # Phone
        }
        
        # Fill each cell
        for cell_addr, value in general_info_data.items():
            sheet[cell_addr] = value
            print(f"  Set General Information!{cell_addr} = {value}")
    else:
        print("Warning: General Information sheet not found")
    
    # Fill Benefit Selection - Group sheet
    print("Filling Benefit Selection - Group sheet...")
    if "Benefit Selection - Group" in wb.sheetnames:
        sheet = wb["Benefit Selection - Group"]
        
        # Define the values to fill (cell_address: value)
        benefit_selection_data = {
            # Category A (Column B)
            'B8': 'Dubai',  # Dubai or AUH policy
            'B9': 'Dubai',  # Region
            'B10': 'Comprehensive Network',  # Select Plan Name
            'B12': 'Abu Dhabi',  # Plan
            'B14': 'GN+',   # Network 
            'B16': 1000000, # AML (Annual Maximum Limit)
            'B18': 'Private',  # IP Room Type
            'B20': 0,       # Coinsurance
            'B22': 'Nil',   # Deductible for Consultation
            'B24': 'Worldwide at R&C of UAE',  # Area of Coverage
            'B26': '0%',    # OP Co-pay for Diagnostics
            'B28': 0,       # Pharmacy Co-Pay
            'B30': 'Upto AML',  # Pharmacy Limit
            
            # Category B (Column G)
            'G8': 'Dubai',
            'G9': 'Dubai',
            'G10': 'Restricted Network',
            'G12': 'Abu Dhabi',
            'G14': 'RN',
            'G16': 250000,
            'G18': 'Shared / Ward',
            'G20': 0,
            'G22': '20% up to AED 50',
            'G24': 'Worldwide at R&C of UAE',
            'G26': '0%',
            'G28': 0.1,     # 10%
            'G30': 'Up to AED 5,000',
            
            # Additional rows (if needed)
            'B32': 'Covered',  # Maternity
            'G32': 'Not Covered',
            'B34': 'Covered',  # Dental
            'G34': 'Not Covered',
            'B36': 'Covered',  # Optical
            'G36': 'Not Covered',
        }
        
        # Fill each cell
        for cell_addr, value in benefit_selection_data.items():
            sheet[cell_addr] = value
            print(f"  Set Benefit Selection - Group!{cell_addr} = {value}")
    else:
        print("Warning: Benefit Selection - Group sheet not found")
    
    # Save the workbook
    print(f"Saving workbook to: {output_path}")
    wb.save(output_path)
    
    print(f"File successfully saved to: {output_path}")
    return output_path


def fill_from_custom_data(file_path, data, output_path=None):
    """
    Fill the Excel file with custom data
    
    Args:
        file_path (str): Path to the Excel file
        data (dict): Dictionary with sheet names as keys and cell_address:value dictionaries as values
        output_path (str, optional): Path to save filled file
        
    Example:
        data = {
            'General Information': {
                'D7': 'Company Name',
                'D9': 'New Business'
            },
            'Benefit Selection - Group': {
                'B8': 'Dubai',
                'G8': 'Dubai'
            }
        }
    """
    if output_path is None:
        # Create output filename with timestamp
        dirname, filename = os.path.split(file_path)
        base, ext = os.path.splitext(filename)
        output_path = os.path.join(dirname, f"{base}_custom_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}")
    
    print(f"Loading workbook: {file_path}")
    
    # Load workbook preserving VBA
    wb = load_workbook(file_path, keep_vba=True)
    
    # Process each sheet in the data dictionary
    for sheet_name, cell_data in data.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in workbook")
            continue
            
        print(f"Filling sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Fill the cells with the provided values
        for cell_addr, value in cell_data.items():
            sheet[cell_addr] = value
            print(f"  Set {sheet_name}!{cell_addr} = {value}")
    
    # Save the workbook
    print(f"Saving workbook to: {output_path}")
    wb.save(output_path)
    
    print(f"File successfully saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    # Example usage with default predefined values
    file_path = "RGA Tool v4.5 - QHX-250425-FALAK TAYYEB PLATINUM GOVERNMENT SERVICES L.L.C-02.04.25.xlsm"
    
    # Method 1: Using predefined values
    filled_file = fill_rga_tool(file_path)
    print(f"File filled with predefined values: {filled_file}")
    
    # Method 2: Using custom data
    custom_data = {
        'General Information': {
            'C5': 'Group',
            'D7': 'ACME CORPORATION',
            'D9': 'New Business',
            'D11': datetime(2024, 5, 1),
            'D13': datetime(2025, 5, 1),
            'D15': datetime(2026, 4, 30),
        },
        'Benefit Selection - Group': {
            'B8': 'Dubai',
            'B14': 'GN+',
            'G14': 'RN',
        }
    }
    
    custom_file = fill_from_custom_data(file_path, custom_data)
    print(f"File filled with custom data: {custom_file}")